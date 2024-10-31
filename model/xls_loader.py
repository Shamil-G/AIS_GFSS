from   db_config import report_db_user, report_db_password, report_db_dsn
import oracledb
from  db.connect import plsql_execute
from  app_config import LOG_PATH
from  util.logger import log
from openpyxl import load_workbook
import datetime
import os.path
import csv


def create_insert_command(t_name, cols: list):
    start = 0
    cmd = f'insert into {t_name}('
    for col_name in cols:
        if start==0:
            cmd = cmd + f'{col_name}'
            start=1
        else:
            cmd = cmd + f', {col_name}'
    cmd = cmd + f') values('
    start=0
    for col_name in cols:
        if start==0:
            cmd = cmd + f':{col_name}'
            start=1
        else:
            cmd = cmd + f', :{col_name}'
    cmd = cmd + f')'
    return len(cols), cmd


# def get_file_name(file_name):
#     file_path = cfg.UPLOAD_PATH + '/' + file_name
#     path = os.path.normpath(file_path)
#     return path

def load_csv(file_name, table_name: str, columns: list):
    global stmt_load
    if not file_name.endswith('.csv'):
        return 0
    
    count_columns, stmt_load = create_insert_command(table_name, columns)
    log.info(f'COUNT_COLUMNS: {count_columns}, STMT: {stmt_load}')

    full_path = file_name
    with open(full_path) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        line_count = 0
        with oracledb.connect(user=report_db_user, password=report_db_password, dsn=report_db_dsn) as connection:
            with connection.cursor() as cursor:
                # CLEAR TABLE BEFORE LOAD
                cursor.execute(f'truncate table {table_name}')
                
                params = []
                for row in csv_reader:
                    if line_count == 0:
                        print(f'Column names are {", ".join(row)}')
                        line_count += 1
                    else:
                        params.clear()
                        print(f'row: {row}')
                        for col in range(1, 3):
                            params.append(row[col])
                        if not params[0]:
                            break
                        log.info(f'-------------------------\n{stmt_load} : {params}')
                        # plsql_execute(cursor, 'load_excel', stmt_load, params)
                        line_count += 1
                cursor.execute('commit')
                print(f'Загружено {line_count-1} строк.')


def load_excel(file_name, table_name: str, columns: list):
    global stmt_load
    cnt_rows: int
    mess = ''

    if not file_name.endswith('.xlsx') and not file_name.endswith('.xls'):
        mess = f"Файл {file_name} не является EXCEL файлом"
        log.info(mess)    
        return 0, mess
    
    count_columns, stmt_load = create_insert_command(table_name, columns)
    log.info(f'COUNT_COLUMNS: {count_columns}, STMT: {stmt_load}')
    
    s_now = datetime.datetime.now()
    file_path = file_name
    path = os.path.normpath(file_path)

    log.info("Загрузка EXCEL стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        mess = f"EXCEL файл не существует. Он должен иметль полный путь: {file_path}"
        log.info(mess)
        return 0, mess

    log_file = open(f'{LOG_PATH}/load_{table_name}.log','w')

    wb = load_workbook(path)
    sheet_number = len(wb.worksheets)
    log.info(f"Книга загружена. Всего листов: {sheet_number}, путь: {path}\nНачинается загрузка таблицы: {table_name}")
    log_file.write(f"Книга {file_name} загружена. Всего листов: {sheet_number}, путь: {path}\nНачинается загрузка таблицы: {table_name}")
    sheet = wb.active

    cnt_rows=0
    with oracledb.connect(user=report_db_user, password=report_db_password, dsn=report_db_dsn) as connection:
        with connection.cursor() as cursor:
            # CLEAR TABLE BEFORE LOAD
            cursor.execute(f'truncate table {table_name}')
            params = []
            current_row = 0
            for sheet in wb.worksheets:
                for i in range(2, sheet.max_row+1):
                    params.clear()
                    if not sheet.cell(row=i, column=1).value:
                        break
                    for col in range(2, len(columns)+2):
                        if type(sheet.cell(row=i, column=col).value) is str:
                            params.append(str.strip(sheet.cell(row=i, column=col).value))
                        else:
                            params.append(sheet.cell(row=i, column=col).value)
                    try:
                        current_row=current_row+1
                        cursor.execute(stmt_load, params)
                        cnt_rows=cnt_rows+1
                        log_file.write(f'{datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")} - INFO  - {stmt_load}, PARAMS: {params}\n')
                    except oracledb.DatabaseError as e:
                        error, = e.args
                        log_file.write(f"{datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S")} - ERROR - {stmt_load}, PARAMS: {params}\n\t\t\t\tLINE {current_row}: {error.message}\n")
                        log.error(f"{stmt_load}, PARAMS: {params}\n\t\t\t\t\t\t LINE {current_row}: {error.message}\n")
                                        
                    # plsql_execute(cursor, 'load_excel', stmt_load, params)
                    # cnt_rows=cnt_rows+1
                now = datetime.datetime.now()
                log.info(f'Загрузка sheet {sheet} завершена.\n+++++ Загружено {cnt_rows}/{sheet.max_row}({current_row}) записей. {now.strftime("%d-%m-%Y %H:%M:%S")}')
                log_file.write(f"{now.strftime("%d-%m-%Y %H:%M:%S")} - INFO  - Загрузка sheet {sheet} завершена.\n+++++ Загружено {cnt_rows}/{sheet.max_row}({current_row}) записей. {now.strftime("%d-%m-%Y %H:%M:%S")}\n")
                log_file.close()                
            cursor.execute('commit')
    return cnt_rows, f'Загружено {cnt_rows} строк'


if __name__ == "__main__":
    log.info(f"TEST")
