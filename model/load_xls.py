from db.connect import get_connection, plsql_execute
import loads_config as cfg
from   openpyxl import load_workbook
from   util.logger import log
import datetime
import os.path
import csv


def create_insert_command():
    start = 0
    cmd = f'insert into {cfg.table}('
    for col_name in cfg.columns:
        if start==0:
            cmd = cmd + f'{col_name}'
            start=1
        else:
            cmd = cmd + f', {col_name}'
    cmd = cmd + f') values('
    start=0
    for col_name in cfg.columns:
        if start==0:
            cmd = cmd + f':{col_name}'
            start=1
        else:
            cmd = cmd + f', :{col_name}'
    cmd = cmd + f')'
    return len(cfg.columns), cmd


def get_file_name(file_name):
    file_path = cfg.UPLOAD_PATH + '/' + file_name
    path = os.path.normpath(file_path)
    return path

def load_csv(file_name):
    global stmt_load
    count_columns, stmt_load = create_insert_command()
    log.info(f'COUNT_COLUMNS: {count_columns}, STMT: {stmt_load}')

    full_path = get_file_name(file_name)
    with open(full_path) as csv_file:
        csv_reader = csv.reader(csv_file, delimiter=';')
        line_count = 0
        params = []
        with get_connection().cursor() as cursor:
            for row in csv_reader:
                if line_count == 0:
                    print(f'Column names are {", ".join(row)}')
                    line_count += 1
                else:
                    params.clear()
                    print(f'row: {row}')
                    for col in range(0, len(cfg.columns)):
                        params.append(row[col])
                    if not params[0]:
                        break
                    log.info(f'{stmt_load} : {params}')
                    plsql_execute(cursor, 'load_excel', stmt_load, params)
                    line_count += 1
            cursor.execute('commit')
        print(f'Загружено {line_count-1} строк.')


def load_excel(file_name):
    global stmt_load
    cnt_rows: int
    count_columns, stmt_load = create_insert_command()
    log.info(f'COUNT_COLUMNS: {count_columns}, STMT: {stmt_load}')
    
    s_now = datetime.datetime.now()
    #if cfg.os == 'unix':
    file_path = cfg.UPLOAD_PATH + '/' + file_name
    #else:
    #   file_path = cfg.UPLOAD_PATH + '\\' + file_name

    # Нормируем путь к файлу по слэшам
    path = os.path.normpath(file_path)

    log.info("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_name + ' : ' + file_path)

    if not os.path.isfile(file_path):
        log.info(f"File not exists: {str(os.path.isfile(file_path))}")
        return file_name

    wb = load_workbook(path)
    sheet_number = len(wb.worksheets)
    log.info(f"Книга загружена. Всего листов: {sheet_number}, путь: {path}")
    sheet = wb.active

    with get_connection() as conn:
        with conn.cursor() as cursor:
            params = []
            cnt_rows=0
            for sheet in wb.worksheets:
                for i in range(1, sheet.max_row+1):
                    params.clear()
                    for col in range(1, count_columns+1):
                        params.append(sheet.cell(row=i, column=col).value)
                    if cfg.check_null and not params[cfg.check_column]:
                        log.info(f'Exit. Параметры при выходе: {params}, check_column: {cfg.check_column} : {params[cfg.check_column]}')
                        break
                    log.info(f'{stmt_load} : {params} : {params[cfg.check_column]}')
                    plsql_execute(cursor, 'load_excel', stmt_load, params)
                    cnt_rows=cnt_rows+1
                now = datetime.datetime.now()
                log.info(f'Загрузка sheet {sheet} завершена. Загружено {cnt_rows}/{sheet.max_row} записей. {now.strftime("%d-%m-%Y %H:%M:%S")}')
            cursor.execute('commit')
    return


if __name__ == "__main__":
    
    if cfg.filename.endswith('.xlsx'):
        load_excel(cfg.filename)
    if cfg.filename.endswith('.csv'):
        load_csv(cfg.filename)
